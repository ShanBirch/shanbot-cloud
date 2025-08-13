import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import calendar


def create_uber_tax_template():
    """Create a comprehensive Uber tax reimbursement spreadsheet template"""

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Remove default sheet and create our custom sheets
    wb.remove(wb.active)

    # Create main tracking sheet
    ws_main = wb.create_sheet("Uber Expense Tracker")

    # Create monthly summary sheet
    ws_summary = wb.create_sheet("Monthly Summary")

    # Create tax summary sheet
    ws_tax = wb.create_sheet("Tax Summary")

    # === MAIN TRACKING SHEET ===
    setup_main_tracking_sheet(ws_main)

    # === MONTHLY SUMMARY SHEET ===
    setup_monthly_summary_sheet(ws_summary)

    # === TAX SUMMARY SHEET ===
    setup_tax_summary_sheet(ws_tax)

    # Save the workbook
    filename = f"Uber_Tax_Reimbursement_Template_{datetime.now().strftime('%Y')}.xlsx"
    wb.save(filename)

    return filename


def setup_main_tracking_sheet(ws):
    """Setup the main expense tracking sheet"""

    # Headers
    headers = [
        "Date", "Time", "Pickup Location", "Dropoff Location",
        "Trip Purpose", "Business/Personal", "Uber Fare", "Tips",
        "Total Cost", "Receipt #", "Notes", "Reimbursable", "Month", "Week"
    ]

    # Set headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092",
                                end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Set column widths
    column_widths = [12, 8, 25, 25, 20, 15, 12, 8, 12, 15, 30, 12, 10, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(
            col)].width = width

    # Add sample data and formulas for the first few rows
    sample_data = [
        ["2024-01-15", "09:30", "Home", "Client Meeting - CBD", "Client Meeting", "Business", 25.50, 3.00,
            "=G2+H2", "UBR001", "Meeting with ABC Corp", "=IF(F2=\"Business\",I2,0)", "=MONTH(A2)", "=WEEKNUM(A2)"],
        ["2024-01-15", "15:45", "CBD", "Airport", "Business Travel", "Business", 45.80, 5.00, "=G3+H3",
            "UBR002", "Flight to Sydney", "=IF(F3=\"Business\",I3,0)", "=MONTH(A3)", "=WEEKNUM(A3)"],
        ["2024-01-16", "19:20", "Restaurant", "Home", "Dinner", "Personal", 18.30, 2.00, "=G4+H4",
            "UBR003", "Personal dinner", "=IF(F4=\"Business\",I4,0)", "=MONTH(A4)", "=WEEKNUM(A4)"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Format specific columns
            if col_idx == 1:  # Date column
                cell.number_format = 'DD/MM/YYYY'
            elif col_idx == 2:  # Time column
                cell.number_format = 'HH:MM'
            elif col_idx in [7, 8, 9, 12]:  # Money columns
                cell.number_format = '$#,##0.00'

            # Add border
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Add data validation for Business/Personal column
    from openpyxl.worksheet.datavalidation import DataValidation

    business_validation = DataValidation(
        type="list", formula1='"Business,Personal"')
    business_validation.add(f'F2:F1000')
    ws.add_data_validation(business_validation)

    # Add totals row
    totals_row = 100  # Row for totals
    ws.cell(row=totals_row, col=6, value="TOTALS:").font = Font(bold=True)
    ws.cell(row=totals_row, col=7,
            value=f"=SUM(G2:G{totals_row-1})").number_format = '$#,##0.00'
    ws.cell(row=totals_row, col=8,
            value=f"=SUM(H2:H{totals_row-1})").number_format = '$#,##0.00'
    ws.cell(row=totals_row, col=9,
            value=f"=SUM(I2:I{totals_row-1})").number_format = '$#,##0.00'
    ws.cell(row=totals_row, col=12,
            value=f"=SUM(L2:L{totals_row-1})").number_format = '$#,##0.00'

    # Make totals row bold
    for col in range(6, 13):
        ws.cell(row=totals_row, col=col).font = Font(bold=True)
        ws.cell(row=totals_row, col=col).fill = PatternFill(
            start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")


def setup_monthly_summary_sheet(ws):
    """Setup the monthly summary sheet"""

    # Title
    ws.cell(row=1, col=1, value="MONTHLY UBER EXPENSE SUMMARY").font = Font(
        size=16, bold=True)
    ws.merge_cells('A1:F1')
    ws.cell(row=1, col=1).alignment = Alignment(horizontal="center")

    # Year selector
    current_year = datetime.now().year
    ws.cell(row=3, col=1, value="Year:").font = Font(bold=True)
    ws.cell(row=3, col=2, value=current_year)

    # Headers for monthly summary
    headers = ["Month", "Total Trips", "Business Trips",
               "Personal Trips", "Total Amount", "Reimbursable Amount"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, col=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092",
                                end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Add months and formulas
    months = [calendar.month_name[i] for i in range(1, 13)]

    for month_num, month_name in enumerate(months, 1):
        row = 5 + month_num

        # Month name
        ws.cell(row=row, col=1, value=month_name)

        # Total trips (count from main sheet)
        ws.cell(row=row, col=2,
                value=f'=COUNTIF(\'Uber Expense Tracker\'.M:M,{month_num})')

        # Business trips
        ws.cell(row=row, col=3,
                value=f'=COUNTIFS(\'Uber Expense Tracker\'.M:M,{month_num},\'Uber Expense Tracker\'.F:F,"Business")')

        # Personal trips
        ws.cell(row=row, col=4,
                value=f'=COUNTIFS(\'Uber Expense Tracker\'.M:M,{month_num},\'Uber Expense Tracker\'.F:F,"Personal")')

        # Total amount
        ws.cell(row=row, col=5,
                value=f'=SUMIF(\'Uber Expense Tracker\'.M:M,{month_num},\'Uber Expense Tracker\'.I:I)')
        ws.cell(row=row, col=5).number_format = '$#,##0.00'

        # Reimbursable amount
        ws.cell(row=row, col=6,
                value=f'=SUMIF(\'Uber Expense Tracker\'.M:M,{month_num},\'Uber Expense Tracker\'.L:L)')
        ws.cell(row=row, col=6).number_format = '$#,##0.00'

    # Add yearly totals
    totals_row = 18
    ws.cell(row=totals_row, col=1, value="YEARLY TOTAL").font = Font(bold=True)
    ws.cell(row=totals_row, col=2, value="=SUM(B6:B17)").font = Font(bold=True)
    ws.cell(row=totals_row, col=3, value="=SUM(C6:C17)").font = Font(bold=True)
    ws.cell(row=totals_row, col=4, value="=SUM(D6:D17)").font = Font(bold=True)
    ws.cell(row=totals_row, col=5, value="=SUM(E6:E17)").font = Font(bold=True)
    ws.cell(row=totals_row, col=5).number_format = '$#,##0.00'
    ws.cell(row=totals_row, col=6, value="=SUM(F6:F17)").font = Font(bold=True)
    ws.cell(row=totals_row, col=6).number_format = '$#,##0.00'

    # Set column widths
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15


def setup_tax_summary_sheet(ws):
    """Setup the tax summary and reimbursement sheet"""

    # Title
    ws.cell(row=1, col=1, value="TAX & REIMBURSEMENT SUMMARY").font = Font(
        size=16, bold=True)
    ws.merge_cells('A1:D1')
    ws.cell(row=1, col=1).alignment = Alignment(horizontal="center")

    # Tax Year
    current_year = datetime.now().year
    ws.cell(row=3, col=1, value="Tax Year:").font = Font(bold=True)
    ws.cell(row=3, col=2, value=current_year)

    # Business expense summary
    ws.cell(row=5, col=1, value="BUSINESS EXPENSE SUMMARY").font = Font(
        size=14, bold=True)

    summary_items = [
        ("Total Business Trips:", "=COUNTIF('Uber Expense Tracker'.F:F,\"Business\")"),
        ("Total Business Amount:",
         "=SUMIF('Uber Expense Tracker'.F:F,\"Business\",'Uber Expense Tracker'.I:I)"),
        ("Average per Business Trip:", "=IF(B7=0,0,B8/B7)"),
        ("", ""),
        ("Total Personal Trips:", "=COUNTIF('Uber Expense Tracker'.F:F,\"Personal\")"),
        ("Total Personal Amount:",
         "=SUMIF('Uber Expense Tracker'.F:F,\"Personal\",'Uber Expense Tracker'.I:I)"),
        ("", ""),
        ("TOTAL REIMBURSABLE:",
         "=SUMIF('Uber Expense Tracker'.F:F,\"Business\",'Uber Expense Tracker'.I:I)"),
    ]

    for i, (label, formula) in enumerate(summary_items, 7):
        ws.cell(row=i, col=1, value=label).font = Font(bold=True)
        if formula:
            cell = ws.cell(row=i, col=2, value=formula)
            if "Amount" in label or "REIMBURSABLE" in label or "Average" in label:
                cell.number_format = '$#,##0.00'
                if "REIMBURSABLE" in label:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(
                        start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Monthly breakdown for tax purposes
    ws.cell(row=16, col=1, value="MONTHLY BUSINESS EXPENSES").font = Font(
        size=14, bold=True)

    # Headers
    ws.cell(row=17, col=1, value="Month").font = Font(bold=True)
    ws.cell(row=17, col=2, value="Business Amount").font = Font(bold=True)
    ws.cell(row=17, col=3, value="Trip Count").font = Font(bold=True)

    # Monthly data
    months = [calendar.month_name[i] for i in range(1, 13)]
    for month_num, month_name in enumerate(months, 1):
        row = 17 + month_num
        ws.cell(row=row, col=1, value=month_name)
        ws.cell(row=row, col=2,
                value=f'=SUMIFS(\'Uber Expense Tracker\'.I:I,\'Uber Expense Tracker\'.M:M,{month_num},\'Uber Expense Tracker\'.F:F,"Business")')
        ws.cell(row=row, col=2).number_format = '$#,##0.00'
        ws.cell(row=row, col=3,
                value=f'=COUNTIFS(\'Uber Expense Tracker\'.M:M,{month_num},\'Uber Expense Tracker\'.F:F,"Business")')

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Add instructions
    ws.cell(row=32, col=1, value="INSTRUCTIONS FOR USE:").font = Font(
        size=12, bold=True)

    instructions = [
        "1. Enter each Uber trip in the 'Uber Expense Tracker' sheet",
        "2. Categorize each trip as 'Business' or 'Personal'",
        "3. The 'Monthly Summary' sheet will automatically calculate totals",
        "4. This 'Tax Summary' sheet shows your reimbursable amounts",
        "5. Keep all Uber receipts as backup documentation",
        "6. Print or save this summary for tax filing",
        "",
        "TIP: For business trips, include purpose and client/meeting details in Notes column"
    ]

    for i, instruction in enumerate(instructions, 33):
        ws.cell(row=i, col=1, value=instruction)


def main():
    """Main function to create the template"""
    print("Creating Uber Tax Reimbursement Template...")

    try:
        filename = create_uber_tax_template()
        print(f"✅ Success! Created: {filename}")
        print("\n📋 Template Features:")
        print("• Main expense tracking with automatic calculations")
        print("• Monthly summary with trip counts and totals")
        print("• Tax summary with reimbursable amounts")
        print("• Business/Personal categorization")
        print("• Pre-filled formulas for easy use")
        print("• Sample data to get you started")
        print("\n💡 Simply replace the sample data with your actual Uber trips!")

        return filename

    except Exception as e:
        print(f"❌ Error creating template: {e}")
        return None


if __name__ == "__main__":
    main()
